#!/usr/bin/env python3
"""
Extract GHL Opportunities and Form Submissions for Sept, Oct, Nov 2025
Combines both datasets into a single Excel file with multiple sheets
"""

import requests
import json
from datetime import datetime, timezone
from collections import defaultdict
import time

# Openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl for Excel export...")
    import subprocess
    subprocess.check_call(['pip3', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# GHL API Configuration
GHL_API_KEY = 'pit-22f8af95-3244-41e7-9a52-22c87b166f5a'
GHL_LOCATION_ID = 'QdLXaFEqrdF0JbVbpKLw'
GHL_BASE_URL = 'https://services.leadconnectorhq.com'

# Pipeline IDs
ANDRIES_PIPELINE_ID = 'XeAGJWRnUGJ5tuhXam2g'
DAVIDE_PIPELINE_ID = 'pTbNvnrXqJc9u1oxir3q'

# Date ranges for Sept, Oct, Nov 2025
DATE_RANGES = {
    'September': {
        'start': '2025-09-01T00:00:00.000Z',
        'end': '2025-09-30T23:59:59.999Z'
    },
    'October': {
        'start': '2025-10-01T00:00:00.000Z',
        'end': '2025-10-31T23:59:59.999Z'
    },
    'November': {
        'start': '2025-11-01T00:00:00.000Z',
        'end': '2025-11-30T23:59:59.999Z'
    }
}

def get_ghl_headers():
    """Get headers for GHL API requests"""
    return {
        'Authorization': f'Bearer {GHL_API_KEY}',
        'Version': '2021-07-28',
        'Content-Type': 'application/json'
    }

def fetch_all_opportunities():
    """Fetch all opportunities from GHL API with pagination"""
    print("\n" + "="*80)
    print("FETCHING OPPORTUNITIES FROM GHL API")
    print("="*80)
    
    url = f'{GHL_BASE_URL}/opportunities/search'
    all_opportunities = []
    page = 1
    max_retries = 3
    
    while True:
        print(f"\n📄 Fetching page {page}...")
        
        retry_count = 0
        success = False
        
        while retry_count < max_retries and not success:
            try:
                response = requests.get(
                    url,
                    headers=get_ghl_headers(),
                    params={
                        'location_id': GHL_LOCATION_ID,
                        'limit': 100,
                        'page': page
                    },
                    timeout=30  # Add timeout
                )
                response.raise_for_status()
                
                data = response.json()
                opportunities = data.get('opportunities', [])
                
                if not opportunities:
                    print(f"✅ No more opportunities found. Stopping at page {page}")
                    return all_opportunities
                
                all_opportunities.extend(opportunities)
                print(f"   Retrieved {len(opportunities)} opportunities (Total: {len(all_opportunities)})")
                
                if len(opportunities) < 100:
                    print(f"✅ Last page reached (less than 100 results)")
                    return all_opportunities
                
                success = True
                page += 1
                time.sleep(1.0)  # Longer sleep to avoid connection issues
                
            except requests.exceptions.RequestException as e:
                retry_count += 1
                if retry_count < max_retries:
                    print(f"   ⚠️ Connection error (attempt {retry_count}/{max_retries}): {e}")
                    print(f"   ⏳ Waiting 3 seconds before retry...")
                    time.sleep(3)
                else:
                    print(f"   ❌ Failed after {max_retries} attempts: {e}")
                    print(f"   ⚠️ Continuing with {len(all_opportunities)} opportunities fetched so far...")
                    return all_opportunities
    
    print(f"\n✅ Total opportunities fetched: {len(all_opportunities)}")
    return all_opportunities

def fetch_form_submissions(start_date, end_date):
    """Fetch form submissions for a specific date range"""
    url = f'{GHL_BASE_URL}/forms/submissions'
    all_submissions = []
    
    print(f"\n📋 Fetching form submissions from {start_date[:10]} to {end_date[:10]}...")
    
    try:
        # GHL Forms API uses startAt and endAt for date filtering
        response = requests.get(
            url,
            headers=get_ghl_headers(),
            params={
                'locationId': GHL_LOCATION_ID,
                'limit': 100,
                'startAt': start_date,
                'endAt': end_date
            }
        )
        response.raise_for_status()
        
        data = response.json()
        submissions = data.get('submissions', [])
        all_submissions.extend(submissions)
        
        print(f"   Retrieved {len(submissions)} form submissions")
        
        time.sleep(0.5)  # Rate limiting
        
    except requests.exceptions.RequestException as e:
        print(f"❌ Error fetching form submissions: {e}")
    
    return all_submissions

def filter_opportunities_by_month(opportunities):
    """Filter opportunities by month based on createdAt date"""
    monthly_opps = {
        'September': [],
        'October': [],
        'November': []
    }
    
    for opp in opportunities:
        created_at = opp.get('createdAt') or opp.get('dateAdded')
        if not created_at:
            continue
        
        try:
            # Parse the date
            dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            
            # Determine which month
            if dt.year == 2025:
                if dt.month == 9:
                    monthly_opps['September'].append(opp)
                elif dt.month == 10:
                    monthly_opps['October'].append(opp)
                elif dt.month == 11:
                    monthly_opps['November'].append(opp)
        except:
            continue
    
    return monthly_opps

def get_pipeline_name(pipeline_id):
    """Get pipeline name from ID"""
    if pipeline_id == ANDRIES_PIPELINE_ID:
        return 'Andries'
    elif pipeline_id == DAVIDE_PIPELINE_ID:
        return 'Davide'
    else:
        return 'Unknown'

def extract_ad_attribution(opp):
    """Extract Facebook ad attribution from opportunity"""
    attributions = opp.get('attributions', [])
    
    ad_id = ''
    campaign_name = ''
    adset_name = ''
    ad_name = ''
    
    # Look for last attribution
    for attr in reversed(attributions):
        if attr.get('isLast'):
            ad_id = attr.get('h_ad_id') or attr.get('utmAdId') or attr.get('adId') or ''
            campaign_name = attr.get('utmSource', '')
            adset_name = attr.get('utmMedium', '')
            ad_name = attr.get('utmCampaign', '')
            break
    
    # If no "isLast" attribution, take the last one
    if not ad_id and attributions:
        attr = attributions[-1]
        ad_id = attr.get('h_ad_id') or attr.get('utmAdId') or attr.get('adId') or ''
        campaign_name = attr.get('utmSource', '')
        adset_name = attr.get('utmMedium', '')
        ad_name = attr.get('utmCampaign', '')
    
    return ad_id, campaign_name, adset_name, ad_name

def extract_form_submission_data(submission):
    """Extract relevant data from form submission"""
    contact_id = submission.get('contactId', '')
    form_id = submission.get('formId', '')
    name = submission.get('name', '')
    email = submission.get('email', '')
    created_at = submission.get('createdAt', '')
    
    # Extract Facebook attribution from 'others' field
    ad_id = ''
    campaign_id = ''
    adset_id = ''
    
    others = submission.get('others', {})
    if 'eventData' in others and 'url_params' in others['eventData']:
        url_params = others['eventData']['url_params']
        ad_id = url_params.get('ad_id', '')
        campaign_id = url_params.get('campaign_id', '')
        adset_id = url_params.get('adset_id', '')
    elif 'lastAttributionSource' in others:
        attr = others['lastAttributionSource']
        ad_id = attr.get('adId', '')
    
    return {
        'contact_id': contact_id,
        'form_id': form_id,
        'name': name,
        'email': email,
        'created_at': created_at,
        'ad_id': ad_id,
        'campaign_id': campaign_id,
        'adset_id': adset_id
    }

def create_excel_workbook(monthly_opps, monthly_forms):
    """Create Excel workbook with all data"""
    print("\n" + "="*80)
    print("CREATING EXCEL WORKBOOK")
    print("="*80)
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Create Opportunities sheets for each month
    for month in ['September', 'October', 'November']:
        print(f"\n📊 Creating '{month} Opportunities' sheet...")
        
        ws = wb.create_sheet(title=f"{month} Opportunities")
        
        # Headers
        headers = [
            'Opportunity ID', 'Name', 'Contact ID', 'Pipeline', 'Stage', 'Status',
            'Monetary Value (R)', 'Created Date', 'Facebook Ad ID', 'Campaign Name',
            'AdSet Name', 'Ad Name'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        opportunities = monthly_opps[month]
        for row_num, opp in enumerate(opportunities, 2):
            ad_id, campaign_name, adset_name, ad_name = extract_ad_attribution(opp)
            
            ws.cell(row=row_num, column=1, value=opp.get('id', ''))
            ws.cell(row=row_num, column=2, value=opp.get('name', ''))
            ws.cell(row=row_num, column=3, value=opp.get('contactId', ''))
            ws.cell(row=row_num, column=4, value=get_pipeline_name(opp.get('pipelineId', '')))
            ws.cell(row=row_num, column=5, value=opp.get('pipelineStageName', ''))
            ws.cell(row=row_num, column=6, value=opp.get('status', ''))
            
            # Monetary value
            monetary_value = opp.get('monetaryValue', 0)
            if monetary_value:
                ws.cell(row=row_num, column=7, value=float(monetary_value))
            else:
                ws.cell(row=row_num, column=7, value=0)
            
            ws.cell(row=row_num, column=8, value=opp.get('createdAt', ''))
            ws.cell(row=row_num, column=9, value=ad_id)
            ws.cell(row=row_num, column=10, value=campaign_name)
            ws.cell(row=row_num, column=11, value=adset_name)
            ws.cell(row=row_num, column=12, value=ad_name)
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        print(f"   ✅ Added {len(opportunities)} opportunities")
    
    # Create Form Submissions sheets for each month
    for month in ['September', 'October', 'November']:
        print(f"\n📋 Creating '{month} Forms' sheet...")
        
        ws = wb.create_sheet(title=f"{month} Forms")
        
        # Headers
        headers = [
            'Contact ID', 'Form ID', 'Name', 'Email', 'Created Date',
            'Facebook Ad ID', 'Campaign ID', 'AdSet ID'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        forms = monthly_forms[month]
        for row_num, form_data in enumerate(forms, 2):
            ws.cell(row=row_num, column=1, value=form_data['contact_id'])
            ws.cell(row=row_num, column=2, value=form_data['form_id'])
            ws.cell(row=row_num, column=3, value=form_data['name'])
            ws.cell(row=row_num, column=4, value=form_data['email'])
            ws.cell(row=row_num, column=5, value=form_data['created_at'])
            ws.cell(row=row_num, column=6, value=form_data['ad_id'])
            ws.cell(row=row_num, column=7, value=form_data['campaign_id'])
            ws.cell(row=row_num, column=8, value=form_data['adset_id'])
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 25
        
        print(f"   ✅ Added {len(forms)} form submissions")
    
    # Create Summary sheet
    print(f"\n📈 Creating 'Summary' sheet...")
    ws = wb.create_sheet(title="Summary", index=0)
    
    # Title
    ws.cell(row=1, column=1, value="GHL Data Export Summary")
    ws.cell(row=1, column=1).font = Font(size=16, bold=True)
    
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Summary table
    ws.cell(row=4, column=1, value="Month").font = header_font
    ws.cell(row=4, column=2, value="Opportunities").font = header_font
    ws.cell(row=4, column=3, value="Form Submissions").font = header_font
    
    row = 5
    for month in ['September', 'October', 'November']:
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=len(monthly_opps[month]))
        ws.cell(row=row, column=3, value=len(monthly_forms[month]))
        row += 1
    
    # Totals
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=sum(len(monthly_opps[m]) for m in monthly_opps)).font = Font(bold=True)
    ws.cell(row=row, column=3, value=sum(len(monthly_forms[m]) for m in monthly_forms)).font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    print(f"   ✅ Summary sheet created")
    
    return wb

def main():
    """Main execution function"""
    print("\n" + "="*80)
    print("GHL DATA EXTRACTION TO EXCEL")
    print("Extracting Opportunities and Form Submissions for Sept, Oct, Nov 2025")
    print("="*80)
    
    # Step 1: Fetch all opportunities
    all_opportunities = fetch_all_opportunities()
    
    # Step 2: Filter opportunities by month
    print("\n" + "="*80)
    print("FILTERING OPPORTUNITIES BY MONTH")
    print("="*80)
    monthly_opps = filter_opportunities_by_month(all_opportunities)
    
    for month, opps in monthly_opps.items():
        print(f"✅ {month}: {len(opps)} opportunities")
    
    # Step 3: Fetch form submissions for each month
    print("\n" + "="*80)
    print("FETCHING FORM SUBMISSIONS")
    print("="*80)
    
    monthly_forms = {
        'September': [],
        'October': [],
        'November': []
    }
    
    for month, date_range in DATE_RANGES.items():
        submissions = fetch_form_submissions(date_range['start'], date_range['end'])
        monthly_forms[month] = [extract_form_submission_data(sub) for sub in submissions]
        print(f"✅ {month}: {len(monthly_forms[month])} form submissions")
    
    # Step 4: Create Excel workbook
    wb = create_excel_workbook(monthly_opps, monthly_forms)
    
    # Step 5: Save Excel file
    import os
    
    # Create excel_export folder if it doesn't exist
    export_folder = 'excel_export'
    os.makedirs(export_folder, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'GHL_Data_Sept_Oct_Nov_{timestamp}.xlsx'
    filepath = os.path.join(export_folder, filename)
    wb.save(filepath)
    
    print("\n" + "="*80)
    print("EXPORT COMPLETE!")
    print("="*80)
    print(f"\n✅ Excel file saved: {filepath}")
    print(f"\nSheets created:")
    print(f"  - Summary")
    for month in ['September', 'October', 'November']:
        print(f"  - {month} Opportunities ({len(monthly_opps[month])} records)")
        print(f"  - {month} Forms ({len(monthly_forms[month])} records)")
    
    print(f"\n📊 Total Records:")
    print(f"  - Opportunities: {sum(len(monthly_opps[m]) for m in monthly_opps)}")
    print(f"  - Form Submissions: {sum(len(monthly_forms[m]) for m in monthly_forms)}")
    
    print("\n" + "="*80)

if __name__ == '__main__':
    main()

