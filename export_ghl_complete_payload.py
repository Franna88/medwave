#!/usr/bin/env python3
"""
Export Complete GHL Payload Data for November 2025
Creates separate Excel files for Opportunities and Form Submissions
with ALL available fields from the API
"""

import requests
import json
from datetime import datetime
import time

# Openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl for Excel export...")
    import subprocess
    subprocess.check_call(['pip3', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# GHL API Configuration
GHL_API_KEY = 'pit-22f8af95-3244-41e7-9a52-22c87b166f5a'
GHL_LOCATION_ID = 'QdLXaFEqrdF0JbVbpKLw'
GHL_BASE_URL = 'https://services.leadconnectorhq.com'

# November 2025 date range
NOVEMBER_START = '2025-11-01T00:00:00.000Z'
NOVEMBER_END = '2025-11-30T23:59:59.999Z'

def get_ghl_headers():
    """Get headers for GHL API requests"""
    return {
        'Authorization': f'Bearer {GHL_API_KEY}',
        'Version': '2021-07-28',
        'Content-Type': 'application/json'
    }

def flatten_dict(d, parent_key='', sep='_'):
    """Flatten nested dictionary for Excel export"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Convert list to JSON string for Excel
            items.append((new_key, json.dumps(v, default=str)))
        else:
            items.append((new_key, v))
    return dict(items)

def fetch_all_november_opportunities():
    """Fetch all opportunities created in November 2025"""
    print("\n" + "="*80)
    print("FETCHING NOVEMBER 2025 OPPORTUNITIES")
    print("="*80)
    
    url = f'{GHL_BASE_URL}/opportunities/search'
    all_opportunities = []
    page = 1
    max_retries = 3
    consecutive_low_november_count = 0  # Track if we're getting very few November opps
    
    while True:
        print(f"\n📄 Fetching page {page}...")
        
        retry_count = 0
        success = False
        
        while retry_count < max_retries and not success:
            try:
                response = requests.get(
                    url,
                    headers=get_ghl_headers(),
                    params={
                        'location_id': GHL_LOCATION_ID,
                        'limit': 100,
                        'page': page
                    },
                    timeout=30
                )
                response.raise_for_status()
                
                data = response.json()
                opportunities = data.get('opportunities', [])
                
                if not opportunities:
                    print(f"✅ No more opportunities found")
                    break
                
                # Filter for November 2025
                november_opps = []
                for opp in opportunities:
                    created_at = opp.get('createdAt') or opp.get('updatedAt')
                    if created_at:
                        try:
                            dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                            if dt.year == 2025 and dt.month == 11:
                                november_opps.append(opp)
                        except:
                            pass
                
                all_opportunities.extend(november_opps)
                print(f"   Retrieved {len(opportunities)} opportunities, {len(november_opps)} from November")
                print(f"   Total November opportunities: {len(all_opportunities)}")
                
                # If we got less than 100, we're done
                if len(opportunities) < 100:
                    print(f"✅ Last page reached")
                    return all_opportunities
                
                # Track if we're getting very few November opportunities (likely past November)
                if len(november_opps) < 10:  # Less than 10% of page is November
                    consecutive_low_november_count += 1
                    if consecutive_low_november_count >= 3:
                        print(f"✅ Very few November opportunities in last 3 pages, likely passed November range")
                        return all_opportunities
                else:
                    consecutive_low_november_count = 0  # Reset counter
                
                success = True
                page += 1
                time.sleep(1.0)
                
            except requests.exceptions.RequestException as e:
                retry_count += 1
                if retry_count < max_retries:
                    print(f"   ⚠️ Connection error (attempt {retry_count}/{max_retries}): {e}")
                    print(f"   ⏳ Waiting 3 seconds before retry...")
                    time.sleep(3)
                else:
                    print(f"   ❌ Failed after {max_retries} attempts")
                    return all_opportunities
    
    print(f"\n✅ Total November opportunities fetched: {len(all_opportunities)}")
    return all_opportunities

def fetch_november_form_submissions():
    """Fetch all form submissions from November 2025"""
    print("\n" + "="*80)
    print("FETCHING NOVEMBER 2025 FORM SUBMISSIONS")
    print("="*80)
    
    url = f'{GHL_BASE_URL}/forms/submissions'
    all_submissions = []
    
    print(f"\n📋 Fetching form submissions from {NOVEMBER_START[:10]} to {NOVEMBER_END[:10]}...")
    
    try:
        response = requests.get(
            url,
            headers=get_ghl_headers(),
            params={
                'locationId': GHL_LOCATION_ID,
                'limit': 100,
                'startAt': NOVEMBER_START,
                'endAt': NOVEMBER_END
            },
            timeout=30
        )
        response.raise_for_status()
        
        data = response.json()
        submissions = data.get('submissions', [])
        all_submissions.extend(submissions)
        
        print(f"   Retrieved {len(submissions)} form submissions")
        
        time.sleep(0.5)
        
    except requests.exceptions.RequestException as e:
        print(f"❌ Error fetching form submissions: {e}")
    
    print(f"\n✅ Total November form submissions fetched: {len(all_submissions)}")
    return all_submissions

def create_opportunities_excel(opportunities):
    """Create Excel file with complete opportunity payload data"""
    print("\n" + "="*80)
    print("CREATING OPPORTUNITIES EXCEL FILE")
    print("="*80)
    
    if not opportunities:
        print("⚠️ No opportunities to export")
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "November Opportunities"
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Flatten all opportunities to get all possible columns
    print("\n📊 Analyzing opportunity structure...")
    flattened_opps = [flatten_dict(opp) for opp in opportunities]
    
    # Get all unique keys across all opportunities
    all_keys = set()
    for opp in flattened_opps:
        all_keys.update(opp.keys())
    
    headers = sorted(list(all_keys))
    print(f"   Found {len(headers)} unique fields")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    print(f"\n📝 Writing {len(opportunities)} opportunities...")
    for row_num, opp_flat in enumerate(flattened_opps, 2):
        for col_num, header in enumerate(headers, 1):
            value = opp_flat.get(header, '')
            # Convert to string if it's a complex type
            if isinstance(value, (dict, list)):
                value = json.dumps(value, default=str)
            ws.cell(row=row_num, column=col_num, value=value)
        
        if row_num % 50 == 0:
            print(f"   Progress: {row_num - 1}/{len(opportunities)} opportunities written")
    
    # Auto-adjust column widths (max 50 chars)
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = min(30, max(15, len(headers[col_num - 1])))
    
    print(f"   ✅ All {len(opportunities)} opportunities written")
    
    return wb

def create_form_submissions_excel(submissions):
    """Create Excel file with complete form submission payload data"""
    print("\n" + "="*80)
    print("CREATING FORM SUBMISSIONS EXCEL FILE")
    print("="*80)
    
    if not submissions:
        print("⚠️ No form submissions to export")
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "November Form Submissions"
    
    # Header styling
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Flatten all submissions to get all possible columns
    print("\n📊 Analyzing form submission structure...")
    flattened_subs = [flatten_dict(sub) for sub in submissions]
    
    # Get all unique keys across all submissions
    all_keys = set()
    for sub in flattened_subs:
        all_keys.update(sub.keys())
    
    headers = sorted(list(all_keys))
    print(f"   Found {len(headers)} unique fields")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    print(f"\n📝 Writing {len(submissions)} form submissions...")
    for row_num, sub_flat in enumerate(flattened_subs, 2):
        for col_num, header in enumerate(headers, 1):
            value = sub_flat.get(header, '')
            # Convert to string if it's a complex type
            if isinstance(value, (dict, list)):
                value = json.dumps(value, default=str)
            ws.cell(row=row_num, column=col_num, value=value)
        
        if row_num % 50 == 0:
            print(f"   Progress: {row_num - 1}/{len(submissions)} submissions written")
    
    # Auto-adjust column widths (max 50 chars)
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = min(30, max(15, len(headers[col_num - 1])))
    
    print(f"   ✅ All {len(submissions)} form submissions written")
    
    return wb

def main():
    """Main execution function"""
    print("\n" + "="*80)
    print("GHL COMPLETE PAYLOAD EXPORT - NOVEMBER 2025")
    print("="*80)
    
    import os
    
    # Create excel_export folder if it doesn't exist
    export_folder = 'excel_export'
    os.makedirs(export_folder, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Fetch and export opportunities
    opportunities = fetch_all_november_opportunities()
    if opportunities:
        wb_opps = create_opportunities_excel(opportunities)
        if wb_opps:
            filename_opps = f'GHL_November_Opportunities_Complete_{timestamp}.xlsx'
            filepath_opps = os.path.join(export_folder, filename_opps)
            wb_opps.save(filepath_opps)
            print(f"\n✅ Opportunities Excel saved: {filepath_opps}")
    
    # Fetch and export form submissions
    submissions = fetch_november_form_submissions()
    if submissions:
        wb_subs = create_form_submissions_excel(submissions)
        if wb_subs:
            filename_subs = f'GHL_November_FormSubmissions_Complete_{timestamp}.xlsx'
            filepath_subs = os.path.join(export_folder, filename_subs)
            wb_subs.save(filepath_subs)
            print(f"\n✅ Form Submissions Excel saved: {filepath_subs}")
    
    print("\n" + "="*80)
    print("EXPORT COMPLETE!")
    print("="*80)
    print(f"\n📊 Summary:")
    print(f"  - Opportunities exported: {len(opportunities)}")
    print(f"  - Form Submissions exported: {len(submissions)}")
    print(f"\n📁 Files saved to: {export_folder}/")
    print("\n" + "="*80)

if __name__ == '__main__':
    main()

