#!/usr/bin/env python3
"""
Export Complete Facebook Ads Payload for November 2025
Creates an Excel file with ALL available fields from Facebook Ads API
"""

import requests
import json
from datetime import datetime
import time

# Openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl for Excel export...")
    import subprocess
    subprocess.check_call(['pip3', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# Facebook API Configuration
FB_ACCESS_TOKEN = "EAAc9pw8rgA0BP0S8U9s2cLzSJbCYmJZBKZCTFUNDD2zVXVqkC45q1BIQaPdZAmtXKbZBk6wjprLclIUUafHJ4icQZAXuuePybTL38pNQIcjQQZCbRGGhAtLcLVSGeJP59nMdpt8KNEoMQtvDfZBwBgpLNhQboPpaaeU8fW2rCEEhZA9pRN4RjZAAnwnLqEDaP8Fueo0cZD"
FB_AD_ACCOUNT_ID = "act_220298027464902"
FB_API_VERSION = "v24.0"

# November 2025 date range
START_DATE = '2025-11-01'
END_DATE = '2025-11-30'

def validate_facebook_token():
    """Validate Facebook Access Token"""
    print('🔑 Validating Facebook Access Token...')
    
    test_url = f'https://graph.facebook.com/{FB_API_VERSION}/me'
    
    try:
        response = requests.get(test_url, params={
            'access_token': FB_ACCESS_TOKEN,
            'fields': 'id,name'
        })
        
        if response.status_code == 200:
            data = response.json()
            print(f'✅ Token valid for: {data.get("name", "Unknown")}\n')
            return True
        else:
            print(f'❌ Token validation failed: {response.status_code}')
            print(f'   Response: {response.text[:200]}\n')
            return False
    except Exception as e:
        print(f'❌ Error validating token: {e}\n')
        return False

def flatten_dict(d, parent_key='', sep='_'):
    """Flatten nested dictionary for Excel export"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Convert list to JSON string for Excel
            items.append((new_key, json.dumps(v, default=str)))
        else:
            items.append((new_key, v))
    return dict(items)

def fetch_november_ads():
    """Fetch all ads that ran in November 2025"""
    print("\n" + "="*80)
    print("FETCHING NOVEMBER 2025 FACEBOOK ADS")
    print("="*80)
    
    url = f'https://graph.facebook.com/{FB_API_VERSION}/{FB_AD_ACCOUNT_ID}/ads'
    
    params = {
        'access_token': FB_ACCESS_TOKEN,
        'fields': ','.join([
            'id',
            'name',
            'status',
            'effective_status',
            'configured_status',
            'created_time',
            'updated_time',
            'account_id',
            'campaign_id',
            'campaign{id,name,status,objective,daily_budget,lifetime_budget}',
            'adset_id',
            'adset{id,name,status,targeting,optimization_goal,billing_event,daily_budget,lifetime_budget,start_time,end_time}',
            'creative{id,name,thumbnail_url,object_story_spec,asset_feed_spec}',
            'tracking_specs',
            'conversion_specs',
            'bid_type',
            'bid_amount'
        ]),
        'time_range': json.dumps({
            'since': START_DATE,
            'until': END_DATE
        }),
        'limit': 100
    }
    
    all_ads = []
    page = 1
    
    while True:
        print(f"\n📄 Fetching page {page}...")
        
        try:
            response = requests.get(url, params=params)
            response.raise_for_status()
            
            data = response.json()
            ads = data.get('data', [])
            
            if not ads:
                print(f"✅ No more ads found")
                break
            
            all_ads.extend(ads)
            print(f"   Retrieved {len(ads)} ads (Total: {len(all_ads)})")
            
            # Check for next page
            paging = data.get('paging', {})
            next_url = paging.get('next')
            
            if not next_url:
                print(f"✅ Last page reached")
                break
            
            # Update URL for next page
            url = next_url
            params = {}  # Next URL already has all params
            page += 1
            time.sleep(0.5)  # Rate limiting
            
        except requests.exceptions.RequestException as e:
            print(f"❌ Error fetching ads: {e}")
            break
    
    print(f"\n✅ Total November ads fetched: {len(all_ads)}")
    return all_ads

def fetch_ad_insights(ad_id):
    """Fetch insights (metrics) for a specific ad"""
    url = f'https://graph.facebook.com/{FB_API_VERSION}/{ad_id}/insights'
    
    params = {
        'access_token': FB_ACCESS_TOKEN,
        'time_range': json.dumps({
            'since': START_DATE,
            'until': END_DATE
        }),
        'fields': ','.join([
            'impressions',
            'clicks',
            'spend',
            'reach',
            'frequency',
            'cpc',
            'cpm',
            'cpp',
            'ctr',
            'actions',
            'action_values',
            'conversions',
            'cost_per_action_type',
            'inline_link_clicks',
            'inline_link_click_ctr',
            'unique_clicks',
            'unique_ctr'
        ])
    }
    
    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        data = response.json()
        insights = data.get('data', [])
        return insights[0] if insights else {}
    except Exception as e:
        return {}

def create_ads_excel(ads_with_insights):
    """Create Excel file with complete Facebook Ads payload data"""
    print("\n" + "="*80)
    print("CREATING FACEBOOK ADS EXCEL FILE")
    print("="*80)
    
    if not ads_with_insights:
        print("⚠️ No ads to export")
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "November 2025 Ads"
    
    # Header styling
    header_fill = PatternFill(start_color='1877F2', end_color='1877F2', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Flatten all ads to get all possible columns
    print("\n📊 Analyzing ad structure...")
    flattened_ads = []
    for ad in ads_with_insights:
        flattened = flatten_dict(ad)
        flattened_ads.append(flattened)
    
    # Get all unique keys across all ads
    all_keys = set()
    for ad in flattened_ads:
        all_keys.update(ad.keys())
    
    headers = sorted(list(all_keys))
    print(f"   Found {len(headers)} unique fields")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    print(f"\n📝 Writing {len(ads_with_insights)} ads...")
    for row_num, ad_flat in enumerate(flattened_ads, 2):
        for col_num, header in enumerate(headers, 1):
            value = ad_flat.get(header, '')
            # Convert to string if it's a complex type
            if isinstance(value, (dict, list)):
                value = json.dumps(value, default=str)
            ws.cell(row=row_num, column=col_num, value=value)
        
        if row_num % 50 == 0:
            print(f"   Progress: {row_num - 1}/{len(ads_with_insights)} ads written")
    
    # Auto-adjust column widths (max 50 chars)
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = min(40, max(15, len(headers[col_num - 1])))
    
    print(f"   ✅ All {len(ads_with_insights)} ads written")
    
    return wb

def main():
    """Main execution function"""
    print("\n" + "="*80)
    print("FACEBOOK ADS COMPLETE PAYLOAD EXPORT - NOVEMBER 2025")
    print("="*80)
    
    # Validate token
    if not validate_facebook_token():
        print("❌ Cannot proceed without valid token")
        return
    
    # Fetch ads
    ads = fetch_november_ads()
    
    if not ads:
        print("\n⚠️ No ads found for November 2025")
        return
    
    # Fetch insights for each ad
    print("\n" + "="*80)
    print("FETCHING INSIGHTS FOR EACH AD")
    print("="*80)
    
    ads_with_insights = []
    for i, ad in enumerate(ads, 1):
        ad_id = ad.get('id')
        ad_name = ad.get('name', 'Unknown')
        
        if i % 10 == 0 or i == 1:
            print(f"\n📊 Fetching insights {i}/{len(ads)}: {ad_name[:50]}...")
        
        insights = fetch_ad_insights(ad_id)
        
        # Combine ad details with insights
        combined = {**ad, 'insights': insights}
        ads_with_insights.append(combined)
        
        time.sleep(0.3)  # Rate limiting
    
    print(f"\n✅ Fetched insights for all {len(ads_with_insights)} ads")
    
    # Create Excel file
    wb = create_ads_excel(ads_with_insights)
    
    if wb:
        import os
        
        # Create excel_export folder if it doesn't exist
        export_folder = 'excel_export'
        os.makedirs(export_folder, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Facebook_Ads_November_2025_Complete_{timestamp}.xlsx'
        filepath = os.path.join(export_folder, filename)
        wb.save(filepath)
        
        print("\n" + "="*80)
        print("EXPORT COMPLETE!")
        print("="*80)
        print(f"\n✅ Excel file saved: {filepath}")
        print(f"\n📊 Summary:")
        print(f"  - Total ads exported: {len(ads_with_insights)}")
        print(f"  - Date range: {START_DATE} to {END_DATE}")
        print(f"  - Fields included: Complete ad details + insights")
        print("\n" + "="*80)

if __name__ == '__main__':
    main()




